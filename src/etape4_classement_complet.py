#!/usr/bin/env python3
"""Couche 4 — le classement COMPLET des marques, et ce qu'il vaut.

Trier 398 marques les unes contre les autres produit un ordre total. Cet
ordre total est faux. Les ecarts medians sont estimes sur 15 a 2 727
produits selon la marque, avec des intervalles de confiance qui se
chevauchent massivement : la marque de rang 120 et celle de rang 260 ne sont
pas separables par les donnees.

Ce script publie donc trois choses, et pas une seule :

  rang        le rang ponctuel, celui que demande la question ;
  rang_min    et rang_max : l'intervalle de rangs compatible avec les
              donnees. Une marque n'est comptee comme certainement meilleure
              qu'une autre que si leurs IC 95 % sont DISJOINTS. Ce critere
              est conservateur — la non-disjonction n'est pas une absence de
              difference — donc les intervalles de rang publies ici sont, au
              pire, trop larges. Ils ne surestiment jamais la certitude ;
  niveaux     un decoupage en blocs contigus mutuellement separes. Sur ce
              jeu il n'en existe qu'UN : les IC forment une chaine continue
              de la premiere a la derniere marque, sans aucun point de
              coupure. Le resultat est publie comme tel, la colonne ne l'est
              pas — une colonne constante n'informe personne.

Entree : sorties/m_toutes_nutriscore.csv, produit par etape4_marques.py.
Aucun recalcul statistique ici, aucune donnee nouvelle.

Ce classement nomme des entreprises reelles. Il porte sur l'ecart a la
mediane de marche de la strate (sous-categorie x espece) sur le Nutri-Score
en score continu, et sur rien d'autre. Il ne dit rien de la halalite, de la
conformite, ni de la qualite gustative ou sanitaire d'aucun produit.
"""

from __future__ import annotations

import sys

import numpy as np
import pandas as pd

from commun import SORTIES, echec, titre

ENTREE = SORTIES / "m_toutes_nutriscore.csv"
PROFILS = SORTIES / "d4_taux_tag_par_marque.csv"
SEUIL_SPECIALISTE = 50.0     # % de produits tagues halal dans le catalogue

# Signal, pas classification. Le nom d'une marque ne prouve rien sur ses
# produits : Petit Navire vend aussi des rillettes de volaille. Ce motif sert
# uniquement a lever une alerte a verifier produit par produit, parce que
# l'exclusion composee de config/perimetre.yaml laisse passer un produit de la
# mer des lors qu'une categorie a motif carne l'accompagne. Un tel residu
# gonfle mecaniquement le haut du classement : dans une strate « panes », le
# poisson bat le poulet sur le Nutri-Score sans qu'aucune marque n'y soit pour
# quelque chose.
MOTIF_MER_MARQUE = (r"navire|mouette|poissonnier|marin|ocean|oceane|"
                    r"peche|peches|saumon|thon|crevette|coquillage|maree")


def intervalles_de_rang(t: pd.DataFrame) -> pd.DataFrame:
    """Rang ponctuel, et bornes de rang par disjonction des IC 95 %."""
    bas = t.ic95_bas.to_numpy()
    haut = t.ic95_haut.to_numpy()
    # certainement meilleure que i : IC entierement sous l'IC de i.
    meilleures = [(haut < bas[i]).sum() for i in range(len(t))]
    pires = [(bas > haut[i]).sum() for i in range(len(t))]
    t = t.copy()
    t["rang"] = range(1, len(t) + 1)
    t["rang_min"] = [1 + m for m in meilleures]
    t["rang_max"] = [len(t) - p for p in pires]
    return t


def niveaux(t: pd.DataFrame) -> pd.Series:
    """Blocs contigus mutuellement separes par disjonction des IC."""
    niveau, courant, plafond = [], 1, float("-inf")
    for bas, haut in zip(t.ic95_bas, t.ic95_haut):
        if bas > plafond:          # separee de TOUT le bloc en cours
            courant += 1 if plafond > float("-inf") else 0
            plafond = haut
        else:
            plafond = max(plafond, haut)
        niveau.append(courant)
    return pd.Series(niveau, index=t.index)


def separabilite(t: pd.DataFrame) -> tuple[int, int]:
    """Nombre de paires separees, sur le total des paires."""
    bas = t.ic95_bas.to_numpy()
    haut = t.ic95_haut.to_numpy()
    separees = sum((haut < bas[i]).sum() for i in range(len(t)))
    total = len(t) * (len(t) - 1) // 2
    return int(separees), total


def profil_halal(t: pd.DataFrame) -> pd.DataFrame:
    """Distingue le specialiste halal de la marque generaliste a gamme halal.

    « gamme_halal = oui » melange Isla Delice, dont tout le catalogue est
    halal, et Carrefour, dont 2 % des produits carnes le sont. Le rang de
    Carrefour est fixe par ses 98 % non halal : le lire comme un rang halal
    est faux. La sortie D4 donne le taux de produits tagues halal par marque,
    ce qui separe les deux profils.

    D4 ne retient que les marques a 5 produits tagues au moins : une marque
    absente y a donc une gamme halal marginale, ce qui est une information,
    pas une donnee manquante.
    """
    t = t.copy()
    if not PROFILS.exists():
        t["profil"] = "non evalue"
        return t
    d4 = pd.read_csv(PROFILS)[["marque_tag", "pct_tague"]]
    t = t.merge(d4, left_on="marque", right_on="marque_tag", how="left")
    t["profil"] = np.where(
        t.gamme_halal != "oui", "hors bras halal",
        np.where(t.pct_tague.isna(), "gamme halal marginale (< 5 produits)",
                 np.where(t.pct_tague >= SEUIL_SPECIALISTE,
                          "specialiste halal", "gamme halal minoritaire")))
    return t.drop(columns=["marque_tag"])


def alerte_mer(t: pd.DataFrame) -> pd.DataFrame:
    """Marques dont le NOM evoque un produit de la mer. A verifier."""
    return t[t.marque.str.contains(MOTIF_MER_MARQUE, regex=True, na=False)]


def markdown(t: pd.DataFrame, alerte: pd.DataFrame,
             prof: pd.DataFrame | None = None) -> str:
    lignes = [
        "# Classement complet des marques — Nutri-Score, a composition egale",
        "",
        "Ecart a la mediane de marche de la strate (sous-categorie x espece).",
        "**Negatif = meilleur que le marche sur le meme type de produit.**",
        "",
        "`rang` est l'ordre ponctuel, et il ne doit pas etre lu seul :",
        "`rang_min`-`rang_max` est l'intervalle de rangs compatible avec les",
        "donnees (bornes posees par disjonction des IC 95 %, critere",
        "conservateur). Il n'existe aucun point de coupure dans ce classement :",
        "les IC forment une chaine continue du rang 1 au rang 398.",
        "",
        "`n` < 30 : ligne decrite, jamais testee (regle des 30).",
        "`profil` distingue le specialiste halal du generaliste a gamme halal.",
        "Le rang d'un generaliste est fixe par son catalogue majoritairement non",
        "halal : il ne dit rien de sa gamme halal.",
        "",
    ] + ([
        "## Ou se situent les marques halal",
        "",
        "Le rang median des 43 marques a gamme halal est "
        f"{int(t[t.gamme_halal == 'oui'].rang.median())} sur "
        f"{len(t)}. Ce chiffre",
        "est trompeur : il melange le specialiste halal et le generaliste qui a",
        "quelques references halal dans un catalogue qui ne l'est pas. Le rang",
        "d'un generaliste est fixe par ses produits non halal.",
        "",
        "| profil de catalogue | marques | rang median | ecart median | rangs |",
        "|:---|---:|---:|---:|:---:|",
    ] + [
        f"| {i} | {int(r.k)} | {r.rang_median:.0f} | {r.ecart_median:+.1f} "
        f"| {int(r.rang_min)}-{int(r.rang_max)} |"
        for i, r in prof.iterrows()
    ] + [
        "",
        "Seule la ligne « specialiste halal » se lit comme un resultat portant",
        "sur le bras halal.",
        "",
    ] if prof is not None and len(prof) else []) + ([
        "## Alerte : residus de produits de la mer en haut de classement",
        "",
        f"{len(alerte)} marques du classement portent un nom evoquant un produit",
        "de la mer, dont plusieurs dans les 20 premiers rangs et sur une seule",
        "strate. L'exclusion composee du perimetre laisse entrer un produit de la",
        "mer des lors qu'une categorie a motif carne l'accompagne. Dans une strate",
        "comme « panes », le poisson bat le poulet sur le Nutri-Score : un tel",
        "residu gonfle le haut du classement sans qu'aucune marque n'y soit pour",
        "quelque chose. **Le haut de ce classement n'est pas publiable avant**",
        "**verification produit par produit de ces marques.**",
        "",
        "Marques concernees : " + ", ".join(
            f"{r.marque} (rang {r.rang}, n={r.n}, {r.strates_couvertes} strate"
            f"{'s' if r.strates_couvertes > 1 else ''})"
            for r in alerte.itertuples()) + ".",
        "",
        "Le nom d'une marque ne prouve rien sur ses produits, et le motif n'est",
        "pas exhaustif : cette liste est un signal a verifier, pas une exclusion.",
        "",
    ] if len(alerte) else []) + [
        "| rang | rangs possibles | marque | n | ecart | IC 95 % | strates | profil |",
        "|---:|:---:|:---|---:|---:|:---:|---:|:---|",
    ]
    for r in t.itertuples():
        lignes.append(
            f"| {r.rang} | {r.rang_min}-{r.rang_max} | {r.marque} "
            f"| {r.n}{'' if r.regle_30 == 'franchie' else ' *'} "
            f"| {r.ecart_median:+.1f} "
            f"| [{r.ic95_bas:+.1f} ; {r.ic95_haut:+.1f}] "
            f"| {r.strates_couvertes} | {r.profil} |")
    lignes += ["", "`*` effectif sous 30 : ligne descriptive, non testable."]
    return "\n".join(lignes) + "\n"


def tableur(t: pd.DataFrame, chemin) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "classement"
    entetes = ["rang", "rang_min", "rang_max", "marque", "n",
               "ecart_median", "ic95_bas", "ic95_haut", "regle_30",
               "strates_couvertes", "gamme_halal", "profil"]
    ws.append(entetes)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for r in t[entetes].itertuples(index=False):
        ws.append(list(r))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(entetes))}{len(t) + 1}"
    for i, nom in enumerate(entetes, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, len(nom) + 2)
    ws.column_dimensions["D"].width = 34

    # Aucune formule dans ce fichier : une formule non verifiee est une
    # affirmation non verifiee.
    notes = wb.create_sheet("lecture")
    for ligne in [
        ["Classement des marques sur l'ecart a la mediane de marche de leur"],
        ["strate (sous-categorie x espece), Nutri-Score en score continu."],
        [""],
        ["ecart_median  negatif = meilleur que le marche sur le meme produit"],
        ["rang_min/max  intervalle de rangs compatible avec les donnees"],
        ["              (bornes par disjonction des IC 95 %, critere"],
        ["              conservateur : l'intervalle n'est jamais trop etroit)"],
        ["regle_30      'sous 30' = ligne decrite, jamais testee"],
        ["gamme_halal   la marque a au moins un produit du bras halal."],
        ["              Ce n'est pas une marque halal."],
        ["profil        part du catalogue tague halal (sortie D4). Le rang"],
        ["              d'un generaliste est fixe par ses produits non halal"],
        ["              et ne dit rien de sa gamme halal."],
        [""],
        ["Ce classement ne dit rien de la halalite ni de la conformite"],
        ["d'aucun produit."],
    ]:
        notes.append(ligne)
    notes.column_dimensions["A"].width = 72
    wb.save(chemin)


def main() -> int:
    if not ENTREE.exists():
        echec(f"{ENTREE} absent. Lancer d'abord src/etape4_marques.py.")
    t = pd.read_csv(ENTREE).sort_values("ecart_median").reset_index(drop=True)
    t = intervalles_de_rang(t)
    t["niveau"] = niveaux(t)

    titre(f"Classement complet : les {len(t)} marques du perimetre")
    sep, total = separabilite(t)
    print("Ecart a la mediane de marche de la strate, Nutri-Score continu.")
    print("Negatif = meilleur que le marche sur le meme type de produit.\n")
    print(f"  Paires de marques effectivement separees : {sep} sur {total} "
          f"({100 * sep / total:.1f} %).")
    n_niv = int(t.niveau.max())
    print(f"  Blocs contigus mutuellement separes : {n_niv}"
          + (" — aucun point de coupure n'existe dans le classement."
             if n_niv == 1 else "."))
    largeur = (t.rang_max - t.rang_min + 1)
    print(f"  Largeur mediane de l'intervalle de rang : {int(largeur.median())} "
          f"rangs sur {len(t)}.")
    print(f"  Marques dont l'effectif franchit 30 : "
          f"{int((t.regle_30 == 'franchie').sum())}.\n")

    t = profil_halal(t)
    prof = None
    h = t[t.gamme_halal == "oui"]
    print(f"  Les {len(h)} marques a gamme halal : rangs {int(h.rang.min())} a "
          f"{int(h.rang.max())}, rang median {int(h.rang.median())}.")
    print(f"  Mediane de leurs ecarts : {h.ecart_median.median():+.1f} contre "
          f"{t[t.gamme_halal == 'non'].ecart_median.median():+.1f} pour les "
          f"{len(t) - len(h)} autres : les deux groupes sont, en mediane, sur "
          f"la\n  mediane de marche de leurs strates.\n")
    if "profil" in t and t.profil.nunique() > 1:
        print(f"  Ce rang median de {int(h.rang.median())} est trompeur : il "
              "melange des profils que D4\n  separe. Detail par profil de "
              "catalogue :\n")
        prof = (h.groupby("profil")
                 .agg(k=("marque", "size"), rang_median=("rang", "median"),
                      ecart_median=("ecart_median", "median"),
                      rang_min=("rang", "min"), rang_max=("rang", "max"))
                 .sort_values("rang_median"))
        print(prof.to_string())
        prof.to_csv(SORTIES / "classement_profils_halal.csv")

        print("\n  Une marque generaliste est classee par son catalogue, "
              "majoritairement non\n  halal : son rang ne dit rien de sa gamme "
              "halal. Seule la ligne\n  « specialiste halal » se lit comme un "
              "resultat sur le bras halal.\n")

    alerte = alerte_mer(t)
    if len(alerte):
        print("  [ALERTE] Marques au nom evoquant un produit de la mer, "
              "presentes dans le\n  classement malgre l'exclusion composee du "
              "perimetre. A verifier produit\n  par produit avant toute "
              "publication du haut de classement :")
        print(alerte[["rang", "marque", "n", "ecart_median",
                      "strates_couvertes"]].to_string(index=False))
        alerte.to_csv(SORTIES / "classement_alerte_mer.csv", index=False)
        print()

    colonnes = ["rang", "rang_min", "rang_max", "marque", "n",
                "ecart_median", "ic95_bas", "ic95_haut", "regle_30",
                "strates_couvertes", "gamme_halal", "profil"]
    t = t[colonnes]
    t.to_csv(SORTIES / "classement_marques_complet.csv", index=False)
    (SORTIES / "classement_marques_complet.md").write_text(
        markdown(t, alerte, prof), encoding="utf-8")
    tableur(t, SORTIES / "classement_marques_complet.xlsx")

    with pd.option_context("display.max_rows", None, "display.width", 200):
        print(t.to_string(index=False))
    print("\nEcrit : sorties/classement_marques_complet.{csv,md,xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
