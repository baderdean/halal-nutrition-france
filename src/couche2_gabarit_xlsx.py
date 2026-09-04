#!/usr/bin/env python3
"""Couche 2 — classeur de saisie du double codage humain.

Produit un .xlsx pret a remplir a partir de sorties/double_codage_a_remplir.csv.

Deux choix de conception qui ne sont pas cosmetiques :

  1. La colonne `bras` N'APPARAIT PAS. Elle dirait au codeur ce que le tag
     Open Food Facts affirme, or c'est exactement ce que son codage doit
     verifier sans le savoir. Un codage qui connait la reponse attendue ne
     mesure plus rien.
  2. Les lignes sont MELANGEES, graine figee. Triees par bras, elles auraient
     forme deux blocs devinables meme sans la colonne.

Usage : python3 src/couche2_gabarit_xlsx.py
"""

from __future__ import annotations

import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from commun import RACINE, SORTIES, echec

GRAINE = 20260904
SOURCE = SORTIES / "double_codage_a_remplir.csv"
CIBLE = RACINE / "donnees_humaines" / "double_codage_a_remplir.xlsx"

POLICE = "Arial"
JAUNE = PatternFill("solid", fgColor="FFF2CC")     # cellules a remplir
GRIS = PatternFill("solid", fgColor="F2F2F2")      # colonnes en lecture seule
ENTETE = PatternFill("solid", fgColor="1F3864")
BORDURE = Border(*[Side(style="thin", color="BFBFBF")] * 4)

ESTAMPILLE = ["oui", "non", "illisible"]
LISIBILITE = ["nette", "partiel", "floue", "trop_petite", "zone_absente"]

COLONNES = [
    ("n", "N°", 6, False),
    ("code", "Code-barres", 16, False),
    ("brands", "Marque", 26, False),
    ("lien", "Photo (cliquer)", 34, False),
    ("h_estampille_halal", "Estampille halal ?", 18, True),
    ("h_certificateur", "Certificateur lu", 30, True),
    ("h_lisibilite", "Lisibilite", 15, True),
    ("h_commentaire", "Commentaire", 34, True),
]

CONSIGNES = [
    ("Double codage humain — lecture des emballages", "titre"),
    ("", None),
    ("Ce fichier sert a mesurer le taux d'erreur de la lecture automatique "
     "des emballages.", None),
    ("Sans lui, aucun chiffre issu du modele de vision n'est publiable.", None),
    ("", None),
    ("Ce que vous devez remplir", "section"),
    ("Les quatre colonnes en JAUNE de l'onglet « Codage ». Les colonnes "
     "grises ne se modifient pas.", None),
    ("Cliquez le lien de la colonne Photo, regardez l'emballage, repondez.", None),
    ("", None),
    ("Regles de codage", "section"),
    ("1. Ne codez QUE ce que vous voyez sur la photo. Pas ce que vous savez "
     "de la marque.", None),
    ("2. Un nom de certificateur ne se devine pas. Si vous ne le lisez pas, "
     "laissez vide.", None),
    ("3. L'absence de porc, un texte en arabe ou une marque connue ne sont "
     "PAS une estampille.", None),
    ("4. « illisible » correctement code vaut mieux qu'une reponse plausible "
     "mais fausse.", None),
    ("5. Ne consultez pas les sorties de la machine avant d'avoir fini. "
     "Le codage doit rester en aveugle.", None),
    ("", None),
    ("Valeurs autorisees", "section"),
    ("Estampille halal ?   oui  |  non  |  illisible", None),
    ("   oui        une mention ou un logo halal explicite est visible", None),
    ("   non        la photo montre la zone utile, aucune mention halal", None),
    ("   illisible  quelque chose y ressemble mais ne se lit pas", None),
    ("", None),
    ("Lisibilite   nette | partiel | floue | trop_petite | zone_absente", None),
    ("   nette         l'estampille est entierement visible et se lit", None),
    ("   partiel       l'estampille est la mais partiellement masquee ou "
     "coupee", None),
    ("   floue         image trop degradee pour trancher", None),
    ("   trop_petite   la zone existe mais est trop petite pour etre lue", None),
    ("   zone_absente  la face photographiee ne montre pas ou figurerait "
     "l'estampille", None),
    ("", None),
    ("Une regle qui a manque au premier codage", "section"),
    ("Si vous etablissez le statut d'un produit AUTREMENT que par la photo "
     "(site du", None),
    ("fabricant, recherche en ligne), codez lisibilite = zone_absente. C'est "
     "ce qui", None),
    ("permet de separer ensuite ce qui est lisible en rayon de ce qui ne "
     "l'est pas.", None),
    ("", None),
    ("Exemple d'une ligne remplie", "section"),
]


def style_entete(ws, ligne: int) -> None:
    for i, (_, libelle, largeur, _) in enumerate(COLONNES, start=1):
        c = ws.cell(row=ligne, column=i, value=libelle)
        c.font = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
        c.fill = ENTETE
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDURE
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[ligne].height = 30


def main() -> int:
    if not SOURCE.exists():
        echec(f"{SOURCE} absent. Lance d'abord :\n"
              "    python3 src/couche2_lecture_image.py --gabarit-seul "
              "--max 200 --taille full")
    df = pd.read_csv(SOURCE, dtype={"code": str})
    # Melange a graine figee : trie par bras, le fichier formerait deux blocs
    # devinables meme sans la colonne bras.
    df = df.sample(frac=1.0, random_state=GRAINE).reset_index(drop=True)

    wb = Workbook()

    # ---------------------------------------------------------- Consignes
    ws = wb.active
    ws.title = "Consignes"
    ws.column_dimensions["A"].width = 100
    ligne = 1
    for texte, genre in CONSIGNES:
        c = ws.cell(row=ligne, column=1, value=texte)
        if genre == "titre":
            c.font = Font(name=POLICE, size=14, bold=True, color="1F3864")
        elif genre == "section":
            c.font = Font(name=POLICE, size=11, bold=True)
        else:
            c.font = Font(name=POLICE, size=10)
        ligne += 1

    # Exemple d'une ligne remplie, au format exact de l'onglet Codage.
    ligne += 1
    style_entete(ws, ligne)
    exemple = ["1", "3245678901234", "Marque Exemple",
               "https://images.openfoodfacts.org/.../front_fr.4.full.jpg",
               "oui", "AVS", "nette", "logo au dos, sous le code-barres"]
    for i, val in enumerate(exemple, start=1):
        c = ws.cell(row=ligne + 1, column=i, value=val)
        c.font = Font(name=POLICE, size=10,
                      italic=True, color="808080")
        c.border = BORDURE
        c.fill = JAUNE if COLONNES[i - 1][3] else GRIS
    ws.cell(row=ligne + 2, column=1,
            value="Ligne d'illustration seulement : elle ne figure pas dans "
                  "l'onglet Codage.").font = Font(name=POLICE, size=9,
                                                  italic=True, color="808080")

    # Aucune formule dans ce classeur, deliberement. L'environnement de
    # generation ne dispose pas d'un LibreOffice fonctionnel (expiration a
    # 300 s meme sur un fichier de six lignes), donc aucun compteur
    # d'avancement n'a pu etre verifie. Un compteur faux serait pire que pas
    # de compteur : on indique comment lire l'avancement sans formule.
    ligne += 4
    for texte, genre in [
        ("Suivre votre avancement", "section"),
        (f"Il y a {len(df)} lignes a coder dans l'onglet « Codage ».", None),
        ("Selectionnez la colonne E entiere : la barre d'etat en bas de la "
         "fenetre affiche", None),
        ("le nombre de cellules non vides. C'est votre compte de lignes "
         "faites.", None),
        ("", None),
        ("Quand vous avez fini", "section"),
        ("Enregistrez en CSV sous donnees_humaines/double_codage.csv, ou "
         "renvoyez le .xlsx :", None),
        ("la conversion est immediate. Le calcul du taux d'erreur exige au "
         "moins 200 lignes codees.", None),
    ]:
        c = ws.cell(row=ligne, column=1, value=texte)
        c.font = (Font(name=POLICE, size=11, bold=True) if genre == "section"
                  else Font(name=POLICE, size=10))
        ligne += 1

    # ------------------------------------------------------------ Codage
    cs = wb.create_sheet("Codage")
    style_entete(cs, 1)
    cs.freeze_panes = "E2"

    for i, row in df.iterrows():
        r = i + 2
        valeurs = [i + 1, row["code"], row.get("brands", ""), "Voir la photo"]
        for j, val in enumerate(valeurs, start=1):
            c = cs.cell(row=r, column=j, value=val)
            c.font = Font(name=POLICE, size=10)
            c.fill = GRIS
            c.border = BORDURE
        lien = cs.cell(row=r, column=4)
        lien.hyperlink = row["image_url"]
        lien.font = Font(name=POLICE, size=10, color="0563C1", underline="single")
        for j in range(5, 9):
            c = cs.cell(row=r, column=j)
            c.fill = JAUNE
            c.border = BORDURE
            c.font = Font(name=POLICE, size=10)

    dv_e = DataValidation(type="list", formula1=f'"{",".join(ESTAMPILLE)}"',
                          allow_blank=True, showErrorMessage=True,
                          errorTitle="Valeur non autorisee",
                          error="Choisir : " + ", ".join(ESTAMPILLE))
    dv_l = DataValidation(type="list", formula1=f'"{",".join(LISIBILITE)}"',
                          allow_blank=True, showErrorMessage=True,
                          errorTitle="Valeur non autorisee",
                          error="Choisir : " + ", ".join(LISIBILITE))
    cs.add_data_validation(dv_e)
    cs.add_data_validation(dv_l)
    dv_e.add(f"E2:E{len(df) + 1}")
    dv_l.add(f"G2:G{len(df) + 1}")

    CIBLE.parent.mkdir(exist_ok=True)
    wb.save(CIBLE)
    print(f"  {len(df)} lignes  ->  {CIBLE}")
    print("  colonnes a remplir : E estampille, F certificateur, "
          "G lisibilite, H commentaire")
    print("  la colonne `bras` est volontairement absente : le codage doit "
          "rester en aveugle")
    return 0


if __name__ == "__main__":
    sys.exit(main())
